import pytest
from django.urls import reverse
from rest_framework import status
from professionals.models import Professional
import openpyxl
from io import BytesIO

@pytest.mark.django_db
class TestExportExcel:
    def test_export_excel_admin_access(self, admin_client):
        """
        Test that admin can export excel and it contains correct headers and rows.
        """
        from datetime import date
        # Create some data
        pf = Professional.objects.create(
            person_type='PF',
            name='João PF',
            cpf='12345678901',
            email='pf@test.com',
            phone='11999999999',
            birth_date=date(1990, 1, 1),
            zip_code='12345678',
            street='Rua PF',
            number='10',
            neighborhood='Bairro PF',
            city='Cidade PF',
            state='SP',
            education='Enfermeiro',
            institution='USP',
            graduation_year=2020,
            council_name='COREN',
            council_number='123',
            experience_years=5
        )

        pj = Professional.objects.create(
            person_type='PJ',
            name='Razão Social PJ',
            company_name='Nome Fantasia PJ',
            cnpj='12345678000199',
            technical_manager_name='Gestor',
            technical_manager_cpf='98765432100',
            email='pj@test.com',
            phone='11888888888',
            birth_date=date(2000, 1, 1), # Data Abertura
            zip_code='87654321',
            street='Av PJ',
            number='20',
            neighborhood='Centro',
            city='Cidade PJ',
            state='RJ',
            education='Médico',
            institution='UNIFESP',
            graduation_year=2010,
            council_name='CRM',
            council_number='456',
            experience_years=15
        )

        url = reverse('professional-export-excel') 
        
        response = admin_client.get(url)
        assert response.status_code == status.HTTP_200_OK
        assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # Load workbook
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # Check headers
        headers = [cell.value for cell in ws[1]]
        expected_headers = [
            'Data Envio', 'Status', 'Tipo', 
            'Nome / Razão Social', 'Nome Fantasia', 
            'CPF', 'CNPJ', 
            'Data Nascimento / Abertura'
        ]
        # Check first few headers match
        assert headers[:8] == expected_headers
        
        # Verify columns count (28 columns as per views.py)
        assert len(headers) == 28
        
        # Verify rows count (header + 2 pros)
        assert ws.max_row == 3
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        # Find PF row
        pf_row = next(r for r in rows if r[3] == pf.name)
        assert pf_row[2] == 'Pessoa Física'
        assert pf_row[5] == pf.cpf
        assert pf_row[6] == '-' 
        
        # Find PJ row
        pj_row = next(r for r in rows if r[3] == pj.name)
        assert pj_row[2] == 'Pessoa Jurídica'
        assert pj_row[5] == '-' 
        assert pj_row[6] == pj.cnpj
        assert pj_row[4] == 'Nome Fantasia PJ'

    def test_export_excel_denied_for_anon(self, client):
        url = reverse('professional-export-excel')
        response = client.get(url)
        assert response.status_code == status.HTTP_401_UNAUTHORIZED
